from flask import Blueprint, render_template, request, redirect, send_file, url_for, flash
from openpyxl import Workbook
import pytz
from .models import Reparticao, Secretaria, Servico, db, Usuario, Tramite, Telefone
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import io
from sqlalchemy import func
main = Blueprint('main', __name__)
brasilia_tz = pytz.timezone('America/Araguaina')
@main.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        senha = request.form['senha']
        user = Usuario.query.filter_by(email=email).first()
        if user and check_password_hash(user.senha, senha):
            login_user(user)
            return redirect(url_for('main.index'))
        else:
            flash("Usuário ou senha inválidos.")
    return render_template('login.html')

@main.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('main.login'))


@main.route('/')
@login_required
def index():
    return render_template('index.html')

@main.route('/tramite')
@login_required
def tramite_get():
    tramite = Tramite.query.all()
    reparticao = Reparticao.query.all()
    servico = Servico.query.all()
    return render_template('tramite.html', tramite=tramite, reparticao=reparticao, servico=servico)

@main.route('/telefone')
@login_required
def telefone_get():
    telefone = Telefone.query.all()
    servico = Servico.query.all()
    return render_template('telefone.html', telefone=telefone, servico=servico)


@main.route('/tramite', methods=['POST'])
@login_required
def tramite():
    origem = request.form['origem']
    destino = request.form['destino']
    hora_usuario = request.form['tempo']
    status = request.form['status']
    servico_id = request.form['servico'] 
    processo = request.form['processo'] 
    brasilia_tz = pytz.timezone('America/Araguaina')
    agora = datetime.now(brasilia_tz)
    
    try:
        hora_usuario = datetime.strptime(hora_usuario, '%H:%M')
        hora_usuario = hora_usuario.replace(year=agora.year, month=agora.month, day=agora.day)
        hora_usuario = brasilia_tz.localize(hora_usuario)
        tempo_decorrido = agora - hora_usuario
        tempo_em_minutos = tempo_decorrido.total_seconds() / 60 
        servico = db.session.query(Servico).get(servico_id)
        if not servico:
            return "Serviço não encontrado", 404
        novo = Tramite(
            origem=origem,
            destino=destino,
            tempo_atendimento=tempo_em_minutos,  
            status=status,
            usuario=current_user,  
            servico=servico,
            processo=processo
        )
        db.session.add(novo)
        db.session.commit()

        return redirect('/tramite')

    except ValueError:
        return "Formato de hora inválido. Use o formato HH:MM."

@main.route('/telefone', methods=['POST'])
@login_required
def telefone():
    hora_usuario = request.form['tempo']
    status = request.form['status']
    servicos = request.form.getlist('servico')  # Supondo que você receba uma lista de IDs de serviços
    brasilia_tz = pytz.timezone('America/Araguaina')
    agora = datetime.now(brasilia_tz)
    
    try:
        # Convertendo hora do usuário para datetime
        hora_usuario = datetime.strptime(hora_usuario, '%H:%M')
        hora_usuario = hora_usuario.replace(year=agora.year, month=agora.month, day=agora.day)
        hora_usuario = brasilia_tz.localize(hora_usuario)
        
        tempo_decorrido = agora - hora_usuario
        tempo_em_minutos = tempo_decorrido.total_seconds() / 60 

        for servico_id in servicos:  # Loop através dos IDs dos serviços
            servico = db.session.query(Servico).get(servico_id)
            
            if not servico:
                continue  # Se o serviço não for encontrado, pula para o próximo
            
            novo = Telefone(
                tempo_atendimento=tempo_em_minutos,
                status=status,
                usuario=current_user,
                servico=servico
            )
            db.session.add(novo)  # Adiciona o novo tramite para cada serviço

        db.session.commit()  # Comita todas as mudanças no banco de dados de uma vez

        return redirect('/telefone')

    except ValueError:
        return "Formato de hora inválido. Use o formato HH:MM."

@main.route('/registrar', methods=['GET', 'POST'])
def registrar():
    if request.method == 'POST':
        nome = request.form['nome']
        email = request.form['email']
        senha = request.form['senha']

        if Usuario.query.filter_by(email=email).first():
            flash("Este e-mail já está em uso.")
            return render_template('registrar.html')

        usuario = Usuario(
            nome=nome,
            email=email,
            senha=generate_password_hash(senha)
        )
        db.session.add(usuario)
        db.session.commit()
        flash("Usuário registrado com sucesso! Faça login.")
        return redirect(url_for('main.login'))

    return render_template('registrar.html')

@main.route('/cadastros', methods=['GET', 'POST'])
@login_required
def cadastros():
    if request.method == 'POST':
        tipo = request.form.get('tipo')

        if tipo == 'secretaria':
            nome = request.form['nome']
            db.session.add(Secretaria(nome=nome))
            db.session.commit()
            flash('Secretaria cadastrada com sucesso!')

        elif tipo == 'servico':
            nome = request.form['nome']
            secretaria_id = request.form['secretaria']
            db.session.add(Servico(nome=nome, secretaria=secretaria_id))
            db.session.commit()
            flash('Serviço cadastrado com sucesso!')

        elif tipo == 'reparticao':
            if not current_user.administrador:
                flash('Acesso negado. Apenas administradores podem cadastrar repartições.')
                return redirect(url_for('main.cadastros'))
            nome = request.form['nome']
            db.session.add(Reparticao(nome=nome))
            db.session.commit()
            flash('Repartição cadastrada com sucesso!')

        return redirect(url_for('main.cadastros'))

    secretarias = Secretaria.query.all()
    return render_template('cadastros.html', secretarias=secretarias)

@main.route('/relatorios/tramite', methods=['GET', 'POST'])
@login_required
def relatorios():
    """
    Gera relatórios consolidados sobre os trâmites, com base nos registros da tabela Tramite.
    O código foi otimizado para executar as consultas uma única vez e corrigir a lógica de filtro de data.
    """
    # Dicionário para armazenar todos os dados do relatório de forma centralizada
    dados_relatorio = {
        "servicos_no_periodo": 0,
        "tramites_por_usuario": [],
        "media_tempo_servico": 0,
        "processos_por_destino": [],
        "tramites_por_servico": [],
        "status_concluido": 0,
        "status_pendente": 0,
        "status_andamento": 0,
        "inicio": request.form.get('inicio'),
        "fim": request.form.get('fim')
    }

    # Executa a lógica apenas se o formulário for enviado (método POST)
    if request.method == 'POST':
        # Define as datas padrão, usando a data atual como padrão para o fim
        inicio_str = dados_relatorio['inicio'] or '2025-01-01'
        fim_str = dados_relatorio['fim'] or datetime.now(brasilia_tz).strftime('%Y-%m-%d')

        # Atualiza o dicionário para que o formulário mantenha os valores
        dados_relatorio['inicio'] = inicio_str
        dados_relatorio['fim'] = fim_str

        # Converte as strings para objetos datetime
        inicio_dt = datetime.strptime(inicio_str, '%Y-%m-%d')
        fim_dt = datetime.strptime(fim_str, '%Y-%m-%d')

        # --- LÓGICA DE DATA CORRIGIDA ---
        # Cria o limite superior exclusivo para o filtro de data
        fim_dt_limite = fim_dt + timedelta(days=1)

        # --- CONSULTAS EXECUTADAS APENAS UMA VEZ ---

        # 1. Total de serviços (trâmites) no período
        total_servicos = db.session.query(func.count(Tramite.id)).filter(
            Tramite.horario >= inicio_dt, 
            Tramite.horario < fim_dt_limite
        ).scalar()
        dados_relatorio['servicos_no_periodo'] = total_servicos or 0
        
        # 2. Trâmites por usuário
        dados_relatorio['tramites_por_usuario'] = db.session.query(
            Usuario.nome,
            func.count(Tramite.id).label('quantidade')
        ).join(Usuario, Tramite.usuario_id == Usuario.id).filter(
            Tramite.horario >= inicio_dt, 
            Tramite.horario < fim_dt_limite
        ).group_by(Usuario.nome).order_by(Usuario.nome).all()

        # 3. Média de tempo de atendimento
        media = db.session.query(func.avg(Tramite.tempo_atendimento)).filter(
            Tramite.horario >= inicio_dt, 
            Tramite.horario < fim_dt_limite
        ).scalar()
        dados_relatorio['media_tempo_servico'] = round(media, 2) if media else 0

        # 4. Processos por destino
        dados_relatorio['processos_por_destino'] = db.session.query(
            Tramite.destino, 
            func.count(Tramite.id).label('quantidade')
        ).filter(
            Tramite.horario >= inicio_dt, 
            Tramite.horario < fim_dt_limite
        ).group_by(Tramite.destino).all()

        # 5. Trâmites por tipo de serviço
        dados_relatorio['tramites_por_servico'] = db.session.query(
            Servico.nome,
            func.count(Tramite.id).label('quantidade')
        ).join(Servico, Tramite.servico_id == Servico.id).filter(
            Tramite.horario >= inicio_dt, 
            Tramite.horario < fim_dt_limite
        ).group_by(Servico.nome).order_by(Servico.nome).all()

        # 6. Contagem de status
        status_counts = db.session.query(
            Tramite.status, 
            func.count(Tramite.id)
        ).filter(
            Tramite.horario >= inicio_dt, 
            Tramite.horario < fim_dt_limite
        ).group_by(Tramite.status).all()
        
        status_map = dict(status_counts)
        dados_relatorio['status_concluido'] = status_map.get('Concluído', 0)
        dados_relatorio['status_pendente'] = status_map.get('Pendente', 0)
        dados_relatorio['status_andamento'] = status_map.get('Em Andamento', 0)

        # --- Lógica de Ação (Excel ou Relatório na Tela) ---
        if request.form.get('acao') == 'excel':
            wb = Workbook()

            # Aba 1: Resumo
            ws1 = wb.active
            ws1.title = "Resumo Geral"
            ws1.append(["Período", f"{inicio_dt.strftime('%d/%m/%Y')} a {fim_dt.strftime('%d/%m/%Y')}"])
            ws1.append(["Total de Serviços", dados_relatorio['servicos_no_periodo']])
            ws1.append(["Média de Tempo (min)", dados_relatorio['media_tempo_servico']])
            
            # Aba 2: Trâmites por Usuário
            ws2 = wb.create_sheet("Trâmites por Usuário")
            ws2.append(["Usuário", "Quantidade"])
            for nome, quantidade in dados_relatorio['tramites_por_usuario']:
                ws2.append([nome, quantidade])

            # Aba 3: Processos por Destino
            ws3 = wb.create_sheet("Processos por Destino")
            ws3.append(["Destino", "Quantidade"])
            for destino, quantidade in dados_relatorio['processos_por_destino']:
                ws3.append([destino, quantidade])

            # Aba 4: Quantidade por Serviço
            ws4 = wb.create_sheet("Quantidade por Serviço")
            ws4.append(["Serviço", "Quantidade"])
            for nome_servico, quantidade in dados_relatorio['tramites_por_servico']:
                ws4.append([nome_servico, quantidade])
            
            # Aba 5: Status dos Serviços
            ws5 = wb.create_sheet("Status dos Serviços")
            ws5.append(["Status", "Quantidade"])
            ws5.append(["Concluído", dados_relatorio['status_concluido']])
            ws5.append(["Pendente", dados_relatorio['status_pendente']])
            ws5.append(["Em Andamento", dados_relatorio['status_andamento']])

            # Salva e envia o arquivo
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output, as_attachment=True, download_name="relatorio_tramites.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return render_template('relatorios.html', **dados_relatorio)


@main.route('/relatorios/telefone', methods=['GET', 'POST'])
@login_required
def relatorios_de_atendimentos_telefonicos():
    """
    Gera relatórios detalhados sobre os atendimentos telefônicos,
    com base nos registros da tabela Telefone.
    Permite a visualização em página web e a exportação para Excel.
    """
    # Inicializa as variáveis para armazenar os dados dos relatórios
    dados_relatorio = {
        "total_atendimentos": 0,
        "atendimentos_por_usuario": [],
        "media_tempo_atendimento": 0,
        "atendimentos_por_servico": [],
        "status_concluido": 0,
        "status_pendente": 0,
        "status_andamento": 0,
        "inicio": request.form.get('inicio'),
        "fim": request.form.get('fim')
    }

    if request.method == 'POST':
        inicio_str = dados_relatorio['inicio'] if dados_relatorio['inicio'] else '2025-01-01'
        fim_str = dados_relatorio['fim'] if dados_relatorio['fim'] else datetime.now(brasilia_tz).strftime('%Y-%m-%d')

        dados_relatorio['inicio'] = inicio_str
        dados_relatorio['fim'] = fim_str

        # Converte as strings de data para objetos datetime para a consulta
        inicio_dt = datetime.strptime(inicio_str, '%Y-%m-%d')
        fim_dt = datetime.strptime(fim_str, '%Y-%m-%d')

        # --- LÓGICA DE DATA CORRIGIDA ---
        # Cria um limite superior exclusivo para a data final, adicionando 1 dia.
        # Isso garante que todos os registros do dia final (até 23:59:59) sejam incluídos.
        fim_dt_limite = fim_dt + timedelta(days=1)

        # --- Consultas ao Banco de Dados com a Lógica Aprimorada ---

        # 1. Total de atendimentos no período
        total = db.session.query(
            func.count(Telefone.id)
        ).filter(
            Telefone.horario >= inicio_dt, 
            Telefone.horario < fim_dt_limite
        ).scalar()
        dados_relatorio['total_atendimentos'] = total or 0

        # 2. Atendimentos por usuário
        dados_relatorio['atendimentos_por_usuario'] = db.session.query(
            Usuario.nome,
            func.count(Telefone.id).label('quantidade')
        ).join(Usuario, Telefone.usuario_id == Usuario.id).filter(
            Telefone.horario >= inicio_dt, 
            Telefone.horario < fim_dt_limite
        ).group_by(Usuario.nome).order_by(Usuario.nome).all()

        # 3. Média de tempo de atendimento
        media = db.session.query(
            func.avg(Telefone.tempo_atendimento)
        ).filter(
            Telefone.horario >= inicio_dt, 
            Telefone.horario < fim_dt_limite
        ).scalar()
        dados_relatorio['media_tempo_atendimento'] = round(media, 2) if media else 0

        # 4. Atendimentos por tipo de serviço
        dados_relatorio['atendimentos_por_servico'] = db.session.query(
            Servico.nome,
            func.count(Telefone.id).label('quantidade')
        ).join(Servico, Telefone.servico_id == Servico.id).filter(
            Telefone.horario >= inicio_dt, 
            Telefone.horario < fim_dt_limite
        ).group_by(Servico.nome).order_by(Servico.nome).all()

        # 5. Contagem de status dos atendimentos
        status_counts = db.session.query(
            Telefone.status,
            func.count(Telefone.id)
        ).filter(
            Telefone.horario >= inicio_dt, 
            Telefone.horario < fim_dt_limite
        ).group_by(Telefone.status).all()

        status_map = dict(status_counts)
        dados_relatorio['status_concluido'] = status_map.get('Concluído', 0)
        dados_relatorio['status_pendente'] = status_map.get('Pendente', 0)
        dados_relatorio['status_andamento'] = status_map.get('Em Andamento', 0)

        # Se a ação for gerar o arquivo Excel
        if request.form.get('acao') == 'excel':
            wb = Workbook()

            # Aba 1: Resumo Geral
            ws1 = wb.active
            ws1.title = "Resumo Geral"
            # O período no relatório continua usando a data que o usuário selecionou (fim_dt)
            ws1.append(["Período do Relatório", f"{inicio_dt.strftime('%d/%m/%Y')} a {fim_dt.strftime('%d/%m/%Y')}"])
            ws1.append([])
            ws1.append(["Indicador", "Valor"])
            ws1.append(["Total de Atendimentos", dados_relatorio['total_atendimentos']])
            ws1.append(["Média de Tempo de Atendimento (min)", dados_relatorio['media_tempo_atendimento']])

            # Aba 2: Status dos Atendimentos
            ws2 = wb.create_sheet("Status dos Atendimentos")
            ws2.append(["Status", "Quantidade"])
            ws2.append(["Concluído", dados_relatorio['status_concluido']])
            ws2.append(["Pendente", dados_relatorio['status_pendente']])
            ws2.append(["Em Andamento", dados_relatorio['status_andamento']])

            # Aba 3: Atendimentos por Usuário
            ws3 = wb.create_sheet("Atendimentos por Usuário")
            ws3.append(["Usuário", "Quantidade de Atendimentos"])
            for nome, quantidade in dados_relatorio['atendimentos_por_usuario']:
                ws3.append([nome, quantidade])

            # Aba 4: Atendimentos por Serviço
            ws4 = wb.create_sheet("Atendimentos por Serviço")
            ws4.append(["Serviço", "Quantidade de Atendimentos"])
            for nome, quantidade in dados_relatorio['atendimentos_por_servico']:
                ws4.append([nome, quantidade])

            # Salva o arquivo em memória e o envia para download
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                as_attachment=True,
                download_name=f"relatorio_atendimentos_{inicio_str}_a_{fim_str}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # Por padrão, renderiza a página HTML com os dados
    return render_template('relatorios-telefone.html', **dados_relatorio)